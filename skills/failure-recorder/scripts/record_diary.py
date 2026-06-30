import os
import argparse
from datetime import datetime
from openpyxl import Workbook, load_workbook

# 記録を保存するファイルの名前（フルパスで指定）
RECORD_FILE = r'C:\Users\Tetsu\.claude\日記保存エクセル\日々の振り返り記録.xlsx'

def main():
    # argparseを使って、コマンドライン（AIからの命令）からデータを受け取る準備をします
    parser = argparse.ArgumentParser(description="日々の振り返りを記録するスキル")
    parser.add_argument("--achievement", default="特になし", help="今日できたこと")
    parser.add_argument("--failure", default="特になし", help="今日失敗したこと")
    parser.add_argument("--learning", default="特になし", help="そこから学んだこと")
    
    args = parser.parse_args()

    # 現在の日付と時間を取得
    now = datetime.now()
    date_str = now.strftime('%Y-%m-%d %H:%M:%S')

    # フォルダが存在しない場合は作成する
    os.makedirs(os.path.dirname(RECORD_FILE), exist_ok=True)

    # Excelファイルが既に存在するか確認する
    if os.path.exists(RECORD_FILE):
        # 存在する場合は、そのファイルを読み込む
        workbook = load_workbook(RECORD_FILE)
        sheet = workbook.active
    else:
        # 存在しない場合は、新しいExcelファイルを作る
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "振り返り記録"
        # 1行目に「見出し（ヘッダー）」を作る
        sheet.append(['記録日時', '今日できたこと', '今日失敗したこと', '学んだこと'])

    # 一番下の行に、ユーザーがAIに伝えた内容を追記する
    sheet.append([date_str, args.achievement, args.failure, args.learning])

    # Excelファイルを保存する
    workbook.save(RECORD_FILE)

    print("振り返りの記録が完了しました！")
    print(f"できたこと: {args.achievement}")
    print(f"失敗したこと: {args.failure}")
    print(f"学んだこと: {args.learning}")

if __name__ == "__main__":
    main()
