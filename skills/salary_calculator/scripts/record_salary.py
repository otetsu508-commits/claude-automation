import argparse
import os
import json
import math
from datetime import datetime
import openpyxl

def main():
    parser = argparse.ArgumentParser(description='バイトの給料を計算して記録・修正・削除します')
    parser.add_argument('--action', choices=['record', 'update', 'delete'], default='record', help='アクション: record(記録), update(修正), delete(削除)')
    parser.add_argument('--date', required=True, help='対象の勤務日 (YYYY-MM-DD)')
    parser.add_argument('--hours', type=float, help='勤務時間 (例: 5.5)')
    parser.add_argument('--wage', type=int, default=1150, help='時給 (円、デフォルトは1150)')
    parser.add_argument('--transport', type=int, default=0, help='交通費 (円、デフォルトは0)')
    parser.add_argument('--output_dir', default='c:/Users/Tetsu/OneDrive/バイト給料', help='出力するExcelファイルの保存先フォルダ')
    parser.add_argument('--output', help='直接ファイルパスを指定する場合（非推奨ですが互換性のため残しています）')
    
    args = parser.parse_args()
    
    try:
        date_obj = datetime.strptime(args.date, '%Y-%m-%d')
        year = date_obj.year
        month = date_obj.month
        
        # 20日締め対応: 21日以降は翌月として扱う
        if date_obj.day >= 21:
            month += 1
            if month > 12:
                month = 1
                year += 1
    except ValueError:
        print(json.dumps({"status": "error", "message": "日付の形式はYYYY-MM-DDで指定してください。"}, ensure_ascii=False))
        return

    # ファイルパスの決定
    if args.output:
        file_path = args.output
    else:
        file_path = os.path.join(args.output_dir, f"{year}年_バイト給料記録.xlsx")
        
    # 記録・修正する場合は hours と wage が必須
    if args.action in ['record', 'update']:
        if args.hours is None or args.wage is None:
            print(json.dumps({"status": "error", "message": "record または update の場合、--hours と --wage は必須です。"}, ensure_ascii=False))
            return
            
        # 勤務時間を30分刻み（0.5単位）に切り捨て
        adjusted_hours = math.floor(args.hours * 2) / 2.0
        
        # 給料の計算 (時給 × 30分単位の時間 + 交通費)
        calculated_salary = int(adjusted_hours * args.wage) + args.transport
        data = [args.date, adjusted_hours, args.wage, args.transport, calculated_salary]

    headers = ["勤務日", "勤務時間", "時給", "交通費", "日給合計"]
    
    # Excelファイルが存在しない場合の処理
    if not os.path.exists(file_path):
        if args.action in ['update', 'delete']:
            print(json.dumps({"status": "error", "message": "Excelファイルが存在しないため、修正・削除はできません。"}, ensure_ascii=False))
            return
            
        os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)
        wb = openpyxl.Workbook()
        # デフォルトのシートを削除
        default_sheet = wb.active
        wb.remove(default_sheet)
    else:
        wb = openpyxl.load_workbook(file_path)
        
    sheet_name = f"{month}月"
    
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        # 月間合計と平均のためのヘッダー
        ws["G1"] = "月間合計日給"
        ws["H1"] = "月間平均勤務時間"
    else:
        ws = wb[sheet_name]
        
    found_row = None
    # 既存のデータから該当日付を探す (ヘッダー行以降から探す)
    # G列やH列があるためA列のデータが存在する最終行から探すループに変更
    max_row_a = len([cell for cell in ws['A'] if cell.value is not None])
    for row_idx in range(max_row_a, 1, -1):
        cell_date = ws.cell(row=row_idx, column=1).value
        # 文字列としての一致を確認
        if cell_date and str(cell_date).strip() == args.date.strip():
            found_row = row_idx
            break

    message = ""
    calculated_salary = 0
    if args.action in ['record', 'update']:
        calculated_salary = int(adjusted_hours * args.wage) + args.transport

    if args.action == 'record':
        # 次の空き行を探す（A列基準）
        next_row = 2
        while ws.cell(row=next_row, column=1).value is not None:
            next_row += 1
        for col_idx, val in enumerate(data, start=1):
            ws.cell(row=next_row, column=col_idx).value = val
            
        message = f"{args.date} の勤務を記録しました。入力時間: {args.hours}時間 -> 計算対象時間: {adjusted_hours}時間, 合計日給: {calculated_salary}円"
        
    elif args.action == 'update':
        if found_row is None:
            print(json.dumps({"status": "error", "message": f"{args.date} の記録が見つかりませんでした。記録がないため修正できません。"}, ensure_ascii=False))
            return
            
        # 該当行のデータを上書き
        for col_idx, val in enumerate(data, start=1):
            ws.cell(row=found_row, column=col_idx).value = val
        message = f"{args.date} の勤務記録を修正しました。修正後入力時間: {args.hours}時間 -> 計算対象時間: {adjusted_hours}時間, 合計日給: {calculated_salary}円"
        
    elif args.action == 'delete':
        if found_row is None:
            print(json.dumps({"status": "error", "message": f"{args.date} の記録が見つかりませんでした。"}, ensure_ascii=False))
            return
            
        ws.delete_rows(found_row)
        message = f"{args.date} の勤務記録を削除しました。"

    # その月の合計と平均を再計算して書き込む
    total_salary = 0
    total_hours = 0
    count_days = 0
    row_idx = 2
    while ws.cell(row=row_idx, column=1).value is not None:
        try:
            h = float(ws.cell(row=row_idx, column=2).value or 0)
            s = int(ws.cell(row=row_idx, column=5).value or 0)
            total_hours += h
            total_salary += s
            count_days += 1
        except ValueError:
            pass
        row_idx += 1
        
    avg_hours = total_hours / count_days if count_days > 0 else 0
    ws["G2"] = total_salary
    ws["H2"] = round(avg_hours, 2)
    
    # -------------------------
    # 一年のまとめシートの更新
    # -------------------------
    summary_sheet_name = "年間まとめ"
    if summary_sheet_name not in wb.sheetnames:
        # 一番左に作成するかどうかの判断。見やすいように先頭に作成。
        ws_summary = wb.create_sheet(title=summary_sheet_name, index=0)
        ws_summary["A1"] = "月"
        ws_summary["B1"] = "合計日給"
        ws_summary["C1"] = "平均勤務時間"
        ws_summary["A14"] = "年間合計"
    else:
        ws_summary = wb[summary_sheet_name]
        
    # 1月から12月までのデータを収集してまとめシートに記載
    yearly_total_salary = 0
    for m in range(1, 13):
        m_sheet_name = f"{m}月"
        m_total_salary = 0
        m_avg_hours = 0
        if m_sheet_name in wb.sheetnames:
            m_ws = wb[m_sheet_name]
            # 月間合計日給はG2、月間平均勤務時間はH2にある想定
            try:
                m_total_salary = int(m_ws["G2"].value or 0)
                m_avg_hours = float(m_ws["H2"].value or 0)
            except (ValueError, TypeError):
                pass
        
        row_idx = m + 1 # 1月は2行目, 2月は3行目...
        ws_summary[f"A{row_idx}"] = f"{m}月"
        ws_summary[f"B{row_idx}"] = m_total_salary
        ws_summary[f"C{row_idx}"] = m_avg_hours
        
        yearly_total_salary += m_total_salary
        
    ws_summary["B14"] = yearly_total_salary
        
    # 列幅の自動調整 (簡易的) - すべてのシートに対して
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter # 列文字を取得
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

    # 保存
    wb.save(file_path)
    
    # 結果を出力 (AIが読み取る用)
    result = {
        "status": "success",
        "message": message,
        "calculated_salary": calculated_salary if args.action != 'delete' else 0,
        "action": args.action,
        "file": file_path
    }
    print(json.dumps(result, ensure_ascii=False))

if __name__ == '__main__':
    main()
